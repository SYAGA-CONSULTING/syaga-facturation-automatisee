#!/usr/bin/env python3
"""
Génération des rapports Clockify pour facturation avec arrondis
- PDF pour tous les clients  
- Excel spécifique pour LAA
- Utilise l'API Reports Clockify
"""

import os
import requests
import json
from datetime import datetime, timedelta
from pathlib import Path
import math

class ClockifyReportGenerator:
    def __init__(self):
        self.config = self._load_config()
        self.api_key = self.config['CLOCKIFY_API_KEY']
        self.workspace_id = self.config['WORKSPACE_ID']
        
        self.headers = {
            "X-Api-Key": self.api_key,
            "Content-Type": "application/json"
        }
        self.base_url = "https://api.clockify.me/api/v1"
        self.reports_url = "https://reports.api.clockify.me/v1"
        
    def _load_config(self):
        """Charger la configuration Clockify depuis le fichier externe"""
        config_path = Path.home() / '.clockify_config'
        if not config_path.exists():
            raise FileNotFoundError(f"Configuration manquante: {config_path}")
        
        config = {}
        with open(config_path, 'r') as f:
            for line in f:
                if '=' in line and not line.startswith('#'):
                    key, value = line.strip().split('=', 1)
                    config[key] = value
        return config
    
    def round_hours_commercial(self, hours):
        """Arrondir à 10 minutes supérieures (règle commerciale SYAGA)"""
        # 10 minutes = 1/6 d'heure
        return math.ceil(hours * 6) / 6
    
    def export_report_pdf(self, client_name, start_date, end_date, output_path):
        """
        Exporter un rapport PDF via l'API Clockify Reports
        """
        print(f"📄 Génération rapport PDF pour {client_name}...")
        
        # Données pour la requête de rapport détaillé
        report_data = {
            "dateRangeStart": start_date.isoformat() + "Z",
            "dateRangeEnd": end_date.isoformat() + "Z",
            "summaryFilter": {
                "groups": ["PROJECT", "USER", "DATE"]
            },
            "detailedFilter": {
                "page": 1,
                "pageSize": 1000,
                "sortColumn": "DATE"
            },
            "exportType": "PDF",
            "projects": self._get_client_projects(client_name)
        }
        
        # Requête pour générer le rapport
        response = requests.post(
            f"{self.reports_url}/workspaces/{self.workspace_id}/reports/detailed",
            headers=self.headers,
            json=report_data
        )
        
        if response.status_code == 200:
            report_data = response.json()
            
            # Calculer les totaux avec arrondis
            total_hours = 0
            for entry in report_data.get('timeentries', []):
                hours = entry['timeInterval']['duration'] / 3600
                rounded_hours = self.round_hours_commercial(hours)
                total_hours += rounded_hours
            
            print(f"  ✅ {len(report_data.get('timeentries', []))} entrées")
            print(f"  ⏱️ Total arrondi: {total_hours:.2f}h")
            
            # Sauvegarder les données pour génération PDF locale si besoin
            return {
                'client': client_name,
                'period': f"{start_date.date()} au {end_date.date()}",
                'total_hours': total_hours,
                'total_amount': total_hours * 110,  # 110€/h par défaut
                'entries': report_data.get('timeentries', [])
            }
        else:
            print(f"❌ Erreur API: {response.status_code}")
            return None
    
    def export_report_csv(self, client_name, start_date, end_date):
        """
        Exporter en CSV via l'API Reports
        """
        print(f"📊 Export CSV pour {client_name}...")
        
        report_data = {
            "dateRangeStart": start_date.isoformat() + "Z",
            "dateRangeEnd": end_date.isoformat() + "Z",
            "detailedFilter": {
                "exportType": "CSV",
                "page": 1,
                "pageSize": 1000
            },
            "projects": self._get_client_projects(client_name)
        }
        
        # URL pour export CSV direct
        response = requests.post(
            f"{self.reports_url}/workspaces/{self.workspace_id}/reports/detailed/csv",
            headers=self.headers,
            json=report_data
        )
        
        if response.status_code == 200:
            # Sauvegarder le CSV
            filename = f"Clockify_{client_name}_{start_date.strftime('%Y%m')}.csv"
            with open(filename, 'wb') as f:
                f.write(response.content)
            print(f"  ✅ CSV sauvé: {filename}")
            return filename
        else:
            print(f"❌ Erreur CSV: {response.status_code}")
            return None
    
    def export_excel_laa(self, start_date, end_date):
        """
        Export Excel spécifique pour LAA avec mise en forme
        """
        print(f"📈 Génération Excel LAA...")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("❌ openpyxl non installé. Installation...")
            os.system("pip3 install openpyxl --break-system-packages")
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        
        # Récupérer les données LAA
        report_data = self._get_detailed_report("LAA", start_date, end_date)
        
        if not report_data:
            return None
        
        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"LAA_{start_date.strftime('%Y_%m')}"
        
        # En-têtes avec style
        headers = ["Date", "Description", "Heures brutes", "Heures arrondies", "Montant HT"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Données
        row = 2
        total_hours_raw = 0
        total_hours_rounded = 0
        
        for entry in report_data.get('timeentries', []):
            date = entry['timeInterval']['start'][:10]
            description = entry.get('description', 'Sans description')[:50]
            hours_raw = entry['timeInterval']['duration'] / 3600
            hours_rounded = self.round_hours_commercial(hours_raw)
            amount = hours_rounded * 110
            
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=f"{hours_raw:.2f}")
            ws.cell(row=row, column=4, value=f"{hours_rounded:.2f}")
            ws.cell(row=row, column=5, value=f"{amount:.2f}€")
            
            total_hours_raw += hours_raw
            total_hours_rounded += hours_rounded
            row += 1
        
        # Ligne de total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"{total_hours_raw:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"{total_hours_rounded:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"{total_hours_rounded * 110:.2f}€").font = Font(bold=True)
        
        # Ajuster les colonnes
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # Sauvegarder
        filename = f"LAA_Clockify_{start_date.strftime('%Y_%m')}.xlsx"
        wb.save(filename)
        print(f"  ✅ Excel LAA sauvé: {filename}")
        
        return {
            'filename': filename,
            'total_hours_raw': total_hours_raw,
            'total_hours_rounded': total_hours_rounded,
            'total_amount': total_hours_rounded * 110
        }
    
    def _get_client_projects(self, client_name):
        """Récupérer les IDs des projets pour un client"""
        response = requests.get(
            f"{self.base_url}/workspaces/{self.workspace_id}/projects",
            headers=self.headers
        )
        
        if response.status_code == 200:
            projects = response.json()
            client_projects = []
            
            for project in projects:
                if client_name.upper() in project['name'].upper():
                    client_projects.append(project['id'])
            
            return client_projects
        return []
    
    def _get_detailed_report(self, client_name, start_date, end_date):
        """Récupérer un rapport détaillé pour un client"""
        data = {
            "dateRangeStart": start_date.isoformat() + "Z",
            "dateRangeEnd": end_date.isoformat() + "Z",
            "detailedFilter": {
                "page": 1,
                "pageSize": 1000,
                "sortColumn": "DATE"
            },
            "projects": self._get_client_projects(client_name)
        }
        
        response = requests.post(
            f"{self.reports_url}/workspaces/{self.workspace_id}/reports/detailed",
            headers=self.headers,
            json=data
        )
        
        if response.status_code == 200:
            return response.json()
        return None

def main():
    """Génération des rapports pour les 9 factures"""
    print("🕐 GÉNÉRATION RAPPORTS CLOCKIFY POUR FACTURATION")
    print("=" * 80)
    
    try:
        generator = ClockifyReportGenerator()
    except FileNotFoundError as e:
        print(f"❌ {e}")
        print("Créez le fichier ~/.clockify_config avec:")
        print("CLOCKIFY_API_KEY=votre_clé_api")
        print("WORKSPACE_ID=votre_workspace_id")
        return
    
    # Période : mois courant (ajustable)
    now = datetime.now()
    start_date = datetime(now.year, now.month, 1)
    end_date = now
    
    print(f"📅 Période: {start_date.date()} au {end_date.date()}")
    
    # Clients avec factures hors-forfait nécessitant justificatifs
    clients_hf = [
        "LAA",      # Format Excel + PDF
        "AXION",    # PDF
        "FARBOS",   # PDF  
        "ART",      # PDF
        "LEFEBVRE", # PDF
        "PETRAS",   # PDF
        "TOUZEAU"   # PDF
    ]
    
    print(f"\n🎯 Génération pour {len(clients_hf)} clients...")
    
    # LAA : Export Excel spécial
    print(f"\n{'='*60}")
    print("LAA - Export Excel avec arrondis")
    print(f"{'='*60}")
    
    laa_result = generator.export_excel_laa(start_date, end_date)
    if laa_result:
        print(f"📈 LAA Excel généré:")
        print(f"  • Heures brutes: {laa_result['total_hours_raw']:.2f}h")
        print(f"  • Heures arrondies: {laa_result['total_hours_rounded']:.2f}h")
        print(f"  • Montant: {laa_result['total_amount']:.2f}€ HT")
    
    # Autres clients : PDF
    print(f"\n{'='*60}")
    print("Autres clients - Rapports PDF")
    print(f"{'='*60}")
    
    for client in clients_hf:
        if client != "LAA":
            print(f"\n📄 {client}:")
            result = generator.export_report_pdf(client, start_date, end_date, f"{client}_rapport.pdf")
            if result:
                print(f"  • Total: {result['total_hours']:.2f}h = {result['total_amount']:.2f}€ HT")
    
    print(f"\n✅ Génération terminée ! Fichiers créés dans le répertoire courant.")
    print(f"\n📌 Instructions:")
    print(f"1. Joindre le fichier Excel à l'email LAA")
    print(f"2. Joindre les PDFs aux autres factures hors-forfait")
    print(f"3. Vérifier les totaux avec les montants facturés")

if __name__ == "__main__":
    main()